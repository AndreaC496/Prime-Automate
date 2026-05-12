from __future__ import annotations
import os
from typing import Optional

# ── SFR / ROM scoring ─────────────────────────────────────────────────────────

_ROM_KEYWORDS: dict[str, float] = {
    "overhead": 0.90,
    "sopra la testa": 0.90,
    "incline": 0.85,
    "panca a 45": 0.85,
    "seduto": 0.82,
    "hip thrust": 0.85,
    "romanian": 0.82,
    "rumeno": 0.82,
    "stacco a una gamba": 0.82,
    "pull-up": 0.80,
    "chin-up": 0.80,
    "squat": 0.80,
    "deadlift": 0.80,
    "stacco": 0.78,
    "good morning": 0.78,
    "bulgarian": 0.80,
    "affondi": 0.78,
    "spider curl": 0.78,
    "pushdown": 0.40,
    "alzate laterali": 0.40,
    "kickback": 0.45,
    "face pull": 0.55,
    "peck fly": 0.50,
}


def estimate_rom(name: str) -> float:
    n = name.lower()
    best = 0.50
    for keyword, score in _ROM_KEYWORDS.items():
        if keyword in n:
            best = max(best, score)
    return round(min(1.0, max(0.0, best)), 4)


def estimate_sfr(name: str, equipment: str, category: str) -> float:
    eq = equipment.lower()
    n = name.lower()

    _isolation_words = {
        "curl", "extension", "fly", "crossover", "kickback", "pushdown",
        "raise", "alzate", "face pull", "adductor", "abductor", "peck",
        "leg curl", "leg extension",
    }
    is_isolation = any(w in n for w in _isolation_words)

    if "macchina" in eq and is_isolation:
        return 0.90
    if ("cavo" in eq or "cavi" in eq) and is_isolation:
        return 0.80
    if "manubr" in eq and is_isolation:
        return 0.75
    if "macchina" in eq:
        return 0.70
    if "bilanciere" in eq:
        return 0.50
    if "nessuna" in eq or "corpo libero" in eq or "corpo-libero" in eq:
        return 0.60
    return 0.60


def normalize_exercise(raw: dict) -> dict:
    def _get(*keys: str) -> str:
        for k in keys:
            v = raw.get(k)
            if v is not None:
                return str(v).strip()
        return ""

    name = _get("name", "nome", "Name", "Esercizio")
    primary = _get("primaryMuscle", "muscolo_primario", "Muscolo Primario", "primary_muscle", "Gruppo Muscolare Primario")
    secondary = _get("secondaryMuscles", "muscoli_secondari", "Muscoli Secondari")
    category = _get("category", "categoria", "Categoria")
    equipment = _get("equipment", "attrezzatura", "Attrezzatura", "Attrezzatura Necessaria")
    home_gym = _get("homeGym", "home_gym", "Home Gym", "homegym").lower() or "no"

    try:
        sfr = float(raw.get("sfr_score") or raw.get("sfr") or estimate_sfr(name, equipment, category))
    except (TypeError, ValueError):
        sfr = estimate_sfr(name, equipment, category)

    try:
        rom = float(raw.get("rom_score") or raw.get("rom") or estimate_rom(name))
    except (TypeError, ValueError):
        rom = estimate_rom(name)

    return {
        "name": name,
        "primaryMuscle": primary,
        "secondaryMuscles": secondary,
        "category": category,
        "equipment": equipment,
        "homeGym": home_gym,
        "sfr_score": round(min(1.0, max(0.0, sfr)), 4),
        "rom_score": round(min(1.0, max(0.0, rom)), 4),
    }


def chunk_text(text: str, chunk_size: int = 300, overlap: int = 50) -> list[str]:
    words = text.split()
    chunks: list[str] = []
    i = 0
    while i < len(words):
        chunk = " ".join(words[i : i + chunk_size])
        if chunk.strip():
            chunks.append(chunk)
        i += chunk_size - overlap
    return chunks


# ── Document extraction ───────────────────────────────────────────────────────

def extract_docx_text(path: str) -> str:
    import docx  # python-docx
    doc = docx.Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_pdf_text(path: str) -> str:
    import fitz  # PyMuPDF
    doc = fitz.open(path)
    return "\n".join(page.get_text() for page in doc)


def load_exercises_from_xlsx(path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    exercises: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        ex = {headers[i]: val for i, val in enumerate(row) if i < len(headers) and headers[i]}
        exercises.append(ex)
    return exercises


# ── ChromaDB ingestion ────────────────────────────────────────────────────────

def ingest(
    xlsx_path: str,
    docx_path: str,
    pdf_path: str,
    chroma_path: str = "db",
) -> tuple[int, int]:
    import chromadb
    from sentence_transformers import SentenceTransformer

    MODEL_NAME = "paraphrase-multilingual-MiniLM-L12-v2"
    model = SentenceTransformer(MODEL_NAME)
    client = chromadb.PersistentClient(path=chroma_path)

    # ── Exercises ──────────────────────────────────────────────────────────────
    try:
        client.delete_collection("exercises")
    except Exception:
        pass
    ex_col = client.create_collection("exercises")

    raw_list = load_exercises_from_xlsx(xlsx_path)
    normalized = [
        normalize_exercise(r)
        for r in raw_list
        if r.get("name") or r.get("nome") or r.get("Name") or r.get("Esercizio")
    ]

    docs = [
        f"{e['name']} | {e['primaryMuscle']} | {e['category']} | {e['equipment']}"
        for e in normalized
    ]
    embeddings = model.encode(docs, show_progress_bar=True).tolist()
    ids = [f"ex_{i}" for i in range(len(normalized))]
    ex_col.add(documents=docs, embeddings=embeddings, ids=ids, metadatas=normalized)

    # ── Knowledge ──────────────────────────────────────────────────────────────
    try:
        client.delete_collection("knowledge")
    except Exception:
        pass
    kn_col = client.create_collection("knowledge")

    knowledge_text = ""
    if os.path.exists(docx_path):
        knowledge_text += extract_docx_text(docx_path) + "\n"
    if os.path.exists(pdf_path):
        knowledge_text += extract_pdf_text(pdf_path) + "\n"

    chunks = chunk_text(knowledge_text, chunk_size=300, overlap=50)
    kn_embeddings = model.encode(chunks, show_progress_bar=True).tolist()
    kn_ids = [f"kn_{i}" for i in range(len(chunks))]
    kn_metadatas = [{"chunk_index": i} for i in range(len(chunks))]
    kn_col.add(documents=chunks, embeddings=kn_embeddings, ids=kn_ids, metadatas=kn_metadatas)

    return len(normalized), len(chunks)
